import openpyxl, os, shutil, tkinter
from openpyxl.utils import range_boundaries
from tkinter import *
from tkinter import font as tkfont

Invoicemaker = True
while Invoicemaker:
    filename = 'Invoice Excel Docs'
    directory = os.path.abspath(filename)
    invoice_data = {}
    # looks at all previous invoices and stores important customer data
    for invoices in os.listdir(directory):
        if not invoices.endswith('.xlsx'):
            break
        invoice = openpyxl.load_workbook(os.path.join(directory, invoices))
        invoice_sheet = invoice.active
        invoice_number = invoice_sheet['B1'].value
        if invoice_number is None:
            continue
        for value in invoice_number:
            if value.isdigit():
                index = invoice_number.index(value)
                invoice_number = invoice_number[index:]
                break
        invoice_num = int(invoice_number)
        name = invoice_sheet['B7'].value
        address = invoice_sheet['B8'].value
        phone_number = invoice_sheet['B9'].value
        invoice_data[name] = {'customer_name': name, 'address': address, 'phone_number': phone_number}
    Invoicemaker = False

new_invoice = openpyxl.load_workbook('Invoice_template.xlsx')
new_invoice_sheet = new_invoice.active
name_list = []
invoice_info = []
InvoiceMaker = Tk()
InvoiceMaker.title("Invoice Maker")
# Set the size of the window
InvoiceMaker.geometry("550x350")


# Define a function for switching the frames
# changes to data input page 1 or 2
def change_to_input():
    print(invoice_data)
    print(len(invoice_data))
    first_n = first_name.get()
    last_n = last_name.get()
    name_list.append(first_n)
    name_list.append(last_n)
    first_name.delete(0, END)
    last_name.delete(0, END)
    if name_list[0] + ' ' + name_list[1] in invoice_data:
        data_input_page1.pack(fill='both', expand=1)
        welcome_page.pack_forget()
    else:
        data_input_page2.pack(fill='both', expand=1)
        welcome_page.pack_forget()


# changes to exit page and prints if user chooses to
def change_to_exit():
    exit_page.pack(fill='both', expand=1)
    print_page.pack_forget()
    vary = var1.get()
    name = name_list[0] + ' ' + name_list[1]
    if vary == 1:
        os.startfile(f'{name} #{name_list[2]}.xlsx', "print")
        os.startfile(f'{name} #{name_list[2]}.xlsx', "print")


# changes to welcome page
def change_to_welcome():
    welcome_page.pack(fill='both', expand=1)
    exit_page.pack_forget()
    invoice_info.clear()
    name_list.clear()


# changes to print page
def change_to_print():
    print_page.pack(fill='both', expand=1)
    if name_list[0] + ' ' + name_list[1] in invoice_data:
        data_input_page1.pack_forget()
        job_name = job_description1.get()
        job_pricey = job_price1.get()
        job_details = job_deet1.get()
        invoice_info.append(job_name)
        invoice_info.append(job_details)
        invoice_info.append(int(job_pricey))
        job_description1.delete(0, END)
        job_price1.delete(0, END)
        job_deet1.delete(0, END)
    else:
        data_input_page2.pack_forget()
        job_name = job_description2.get()
        job_pricey = job_price2.get()
        job_details = job_deet2.get()
        address = cust_address2.get()
        phone_number = cust_phone2.get()
        invoice_info.append(job_name)
        invoice_info.append(job_details)
        invoice_info.append(int(job_pricey))
        invoice_info.append(address)
        invoice_info.append(phone_number)
        job_description2.delete(0, END)
        job_price2.delete(0, END)
        job_deet2.delete(0, END)
        cust_address2.delete(0, END)
        cust_phone2.delete(0, END)
    save_invoice()


# saves invoice
def save_invoice():
    name = name_list[0] + ' ' + name_list[1]
    # if customer is not new, then it will use the customer data from the previous invoice
    if name in invoice_data:
        new_invoice_num = new_invoice_numy()
        new_invoice_sheet['B1'] = new_invoice_num[1]
        new_invoice_sheet['B7'] = invoice_data[name]['customer_name']
        new_invoice_sheet['B8'] = invoice_data[name]['address']
        new_invoice_sheet['B9'] = invoice_data[name]['phone_number']
        ch = ','
        if ch in invoice_info[1]:
            parts_list = [part.strip() for part in invoice_info[1].split(",")]
            price_list = [part.strip() for part in invoice_info[2].split(",")]
            new_invoice_sheet['C7'] = invoice_info[0]  # job name
            new_invoice_sheet['B13'] = parts_list[0]  # job details
            new_invoice_sheet['C13'] = price_list[0]  # job price
            for deet, price in zip(parts_list[1:], price_list[1:]):
                new_invoice_sheet.table.add_rows(1)
                table_ref = new_invoice_sheet.table.tab.ref
                last_row_index = int(range_boundaries(table_ref)[3])
                new_cell_label = new_invoice_sheet.cell(row=last_row_index, column=1)
                new_cell_amt = new_invoice_sheet.cell(row=last_row_index, column=2)
                new_cell_label.value = deet
                new_cell_amt.value = price
        else:
            new_invoice_sheet['C7'] = invoice_info[0]  # job name
            new_invoice_sheet['B13'] = invoice_info[1]  # job details
            new_invoice_sheet['C13'] = invoice_info[2]  # job price

    # if customer is new, then it will ask for customer data
    else:
        new_invoice_sheet['B7'] = name
        address = invoice_info[3]
        parts = address.split(", ")
        street = parts[0]
        city_state_zip = parts[1:]
        city, state, zip_code = city_state_zip[0], city_state_zip[1], city_state_zip[2]
        new_address = f"{street}\n{city}, {state.upper()} {zip_code}"
        new_invoice_sheet['B8'] = new_address
        new_invoice_num = new_invoice_numy()
        new_invoice_sheet['B1'] = new_invoice_num[1]
        new_invoice_sheet['B9'] = invoice_info[4]  # phone number
        invoice_data[name] = {'customer_name': name, 'address': new_address,
                              'phone_number': new_invoice_sheet['B9'].value}

        # asks for invoice details
        new_invoice_sheet['C7'] = invoice_info[0]  # job name
        new_invoice_sheet['B13'] = invoice_info[1]  # job details
        new_invoice_sheet['C13'] = invoice_info[2]  # job price

    # saves new invoice separately and saves invoice template with updated invoice number
    new_invoice_template = new_invoice
    new_invoice_template.save('Invoice_template.xlsx')
    new_invoice.save(f'{name} #{new_invoice_num[0]}.xlsx')


def new_invoice_numy():
    invoice_num_data = []
    past_invoice_num = new_invoice_sheet['B1'].value
    for num in past_invoice_num:
        if num.isdigit():
            index_v = past_invoice_num.index(num)
            break
    invoice_value = past_invoice_num[index_v:]
    invoice_num = int(invoice_value) + 1
    invoice_num_data.append(invoice_num)
    name_list.append(invoice_num)
    str_num = past_invoice_num[:index_v] + str(invoice_num)
    invoice_num_data.append(str_num)
    return invoice_num_data


# Create fonts for making difference in the frame
font1 = tkfont.Font(family='Helvetica', size='22', weight='bold')
font2 = tkfont.Font(family='Helvetica', size='12')

# make frames
welcome_page = Frame(InvoiceMaker)
data_input_page1 = Frame(InvoiceMaker)
data_input_page2 = Frame(InvoiceMaker)
print_page = Frame(InvoiceMaker)
exit_page = Frame(InvoiceMaker)

# welcome frame widgets
welcome_page.pack(fill='both', expand=1)
wel_label = Label(welcome_page, text="Welcome to the Invoice Maker!", font=font1)
wel_label.pack(pady=20)
name_label = Label(welcome_page, text="Please enter the customer's first and last name:", foreground="blue", font=font2)
name_label.pack(pady=20)
# welcome frame entry boxes
f_name = Label(welcome_page, text='First Name:', font=font2)
f_name.place(x=145, y=150)
first_name = Entry(welcome_page, width=25)
first_name.place(x=245, y=155)
l_name = Label(welcome_page, text='Last Name:', font=font2)
l_name.place(x=145, y=180)
last_name = Entry(welcome_page, width=25)
last_name.place(x=245, y=185)
# welcome frame button
wel_submit = Button(welcome_page, text="Submit", font=font2, command=change_to_input)
wel_submit.pack(side=BOTTOM, pady=25)

# data input frame 1 widgets
data_label1 = Label(data_input_page1, text="Data Input!", font=font1)
data_label1.pack(pady=10)
input_label1 = Label(data_input_page1, text="Input Customer's Information:", font=font2)
input_label1.pack(pady=10)
cust_label1 = Label(data_input_page1, text="Customer Data Was Found!", foreground="green", font=font2)
cust_label1.pack(pady=10)
# data input frame 1 entry boxes
job_label1 = Label(data_input_page1, text="Job Name: ", font=font2)
job_label1.place(x=145, y=180)
job_description1 = Entry(data_input_page1, width=25)
job_description1.place(x=245, y=185)
price_label1 = Label(data_input_page1, text="Job Price: ", font=font2)
price_label1.place(x=145, y=210)
job_price1 = Entry(data_input_page1, width=25)
job_price1.place(x=245, y=215)
deet_label1 = Label(data_input_page1, text="Job Details: ", font=font2)
deet_label1.place(x=145, y=240)
job_deet1 = Entry(data_input_page1, width=25)
job_deet1.place(x=245, y=245)
# data input frame 1 button
data_submit1 = Button(data_input_page1, text="Submit", font=font2, command=change_to_print)
data_submit1.pack(side=BOTTOM, pady=25)

# data input frame 2 widgets
data_label2 = Label(data_input_page2, text="Data Input!", font=font1)
data_label2.pack(pady=10)
input_label2 = Label(data_input_page2, text="Input Customer's Information:", font=font2)
input_label2.pack(pady=10)
cust_label2 = Label(data_input_page2, text="Customer Data Was Not Found!", foreground="red", font=font2)
cust_label2.pack(pady=10)
# data input frame 2 entry boxes
job_label2 = Label(data_input_page2, text="Job Name: ", font=font2)
job_label2.place(x=20, y=180)
job_description2 = Entry(data_input_page2, width=25)
job_description2.place(x=120, y=185)
deet_label2 = Label(data_input_page2, text="Job Details: ", font=font2)
deet_label2.place(x=20, y=210)
job_deet2 = Entry(data_input_page2, width=25)
job_deet2.place(x=120, y=215)
price_label2 = Label(data_input_page2, text="Job price: ", font=font2)
price_label2.place(x=145, y=250)
job_price2 = Entry(data_input_page2, width=25)
job_price2.place(x=245, y=255)
cust_address_label2 = Label(data_input_page2, text="Address: ", font=font2)
cust_address_label2.place(x=285, y=180)
cust_address2 = Entry(data_input_page2, width=25)
cust_address2.place(x=370, y=185)
cust_phone_label2 = Label(data_input_page2, text="Phone #: ", font=font2)
cust_phone_label2.place(x=285, y=210)
cust_phone2 = Entry(data_input_page2, width=25)
cust_phone2.place(x=370, y=215)
data_submit2 = Button(data_input_page2, text="Submit", font=font2, command=change_to_print)
data_submit2.pack(side=BOTTOM, pady=25)

# print frame widgets
PrintInvoice = Label(print_page, text="Would you like to print?", font=font1)
PrintInvoice.pack(pady=10)
var1 = IntVar()
var2 = IntVar()
c1 = Checkbutton(print_page, text='Yes', font=font2, onvalue=1, offvalue=0, height=2, width=10, variable=var1)
c1.pack()
c2 = Checkbutton(print_page, text='No', font=font2, onvalue=1, offvalue=0, height=2, width=10, variable=var2)
c2.pack()
print_submit = Button(print_page, text="Submit", font=font2, command=change_to_exit)
print_submit.pack(side=BOTTOM, pady=25)

# exit frame widgets
exit_label = Label(exit_page, text="Thank you for using Invoice Maker!", font=font1)
exit_label.pack(pady=10)
again_label = Label(exit_page, text="Would you like to make another invoice?", font=font2)
again_label.pack(pady=10)
# exit frame buttons
start_over_b = Button(exit_page, text="New Invoice", font=font2, command=change_to_welcome)
start_over_b.pack(pady=20)
exit_b = Button(exit_page, text="Exit", foreground='red', font=font2, command=quit)
exit_b.pack(pady=20)

InvoiceMaker.mainloop()